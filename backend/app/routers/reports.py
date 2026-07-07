from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse, HTMLResponse
from sqlalchemy.orm import Session
from backend.app.database.session import get_db
from backend.app.models.db_models import Product, Invoice, Business, Order, Report
from backend.app.utils.auth import get_current_user, User
import io
import csv
import datetime
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(prefix="/api/reports", tags=["Reports"])

@router.get("/")
def get_reports(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    business = db.query(Business).filter(Business.user_id == current_user.id).first()
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
        
    return db.query(Report).filter(Report.business_id == business.id).order_by(Report.id.desc()).all()

@router.get("/export/csv")
def export_inventory_csv(report_type: str = "inventory", current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    business = db.query(Business).filter(Business.user_id == current_user.id).first()
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
        
    output = io.StringIO()
    writer = csv.writer(output)
    
    if report_type == "inventory":
        # Write headers
        writer.writerow(["SKU", "Product Name", "Category", "Cost ($)", "Price ($)", "Stock Level", "Safety Threshold", "Last Updated"])
        
        products = db.query(Product).filter(Product.business_id == business.id).all()
        for p in products:
            stock = p.inventory.stock_level if p.inventory else 0
            safety = p.inventory.safety_threshold if p.inventory else 10
            updated = p.inventory.last_updated.strftime("%Y-%m-%d %H:%M:%S") if p.inventory else "N/A"
            writer.writerow([p.sku, p.name, p.category, p.cost, p.price, stock, safety, updated])
            
        filename = f"inventory_report_{datetime.date.today().isoformat()}.csv"
        
    elif report_type == "sales":
        # Write headers
        writer.writerow(["Invoice Number", "Customer Name", "Amount ($)", "Status", "Issue Date", "Due Date"])
        
        invoices = db.query(Invoice).filter(Invoice.business_id == business.id).all()
        for inv in invoices:
            cust_name = inv.customer_id # Simplification: could resolve customer name
            writer.writerow([inv.invoice_number, f"Customer #{inv.customer_id}", inv.amount, inv.status, inv.issue_date.strftime("%Y-%m-%d"), inv.due_date.strftime("%Y-%m-%d") if inv.due_date else "N/A"])
            
        filename = f"sales_report_{datetime.date.today().isoformat()}.csv"
    else:
        raise HTTPException(status_code=400, detail="Invalid report type")
        
    # Create the StreamingResponse
    output.seek(0)
    response = StreamingResponse(io.BytesIO(output.getvalue().encode('utf-8')), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

@router.get("/export/pdf", response_class=HTMLResponse)
def export_pdf_report(report_type: str = "inventory", current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    business = db.query(Business).filter(Business.user_id == current_user.id).first()
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
        
    products = db.query(Product).filter(Product.business_id == business.id).all()
    
    # We return a beautiful HTML template which the browser print dialog or frontend can convert to PDF, or serves as a gorgeous print report!
    # This is highly professional and visual!
    
    rows_html = ""
    total_val = 0.0
    total_stock = 0
    
    for p in products:
        stock = p.inventory.stock_level if p.inventory else 0
        safety = p.inventory.safety_threshold if p.inventory else 10
        total_val += (stock * p.cost)
        total_stock += stock
        status_badge = "<span style='color: #ef4444; font-weight: bold;'>LOW STOCK</span>" if stock <= safety else "<span style='color: #10b981;'>HEALTHY</span>"
        
        rows_html += f"""
        <tr>
            <td style="padding: 12px; border-bottom: 1px solid #2d2d2d; color: #e2e8f0;">{p.sku}</td>
            <td style="padding: 12px; border-bottom: 1px solid #2d2d2d; color: #e2e8f0; font-weight: 500;">{p.name}</td>
            <td style="padding: 12px; border-bottom: 1px solid #2d2d2d; color: #e2e8f0;">{p.category or 'N/A'}</td>
            <td style="padding: 12px; border-bottom: 1px solid #2d2d2d; color: #e2e8f0; text-align: right;">${p.cost:.2f}</td>
            <td style="padding: 12px; border-bottom: 1px solid #2d2d2d; color: #e2e8f0; text-align: right;">${p.price:.2f}</td>
            <td style="padding: 12px; border-bottom: 1px solid #2d2d2d; color: #e2e8f0; text-align: center;">{stock}</td>
            <td style="padding: 12px; border-bottom: 1px solid #2d2d2d; color: #e2e8f0; text-align: center;">{safety}</td>
            <td style="padding: 12px; border-bottom: 1px solid #2d2d2d; text-align: center;">{status_badge}</td>
        </tr>
        """
        
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>BizPilot AI OS - {report_type.capitalize()} Report</title>
        <style>
            body {{
                font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
                background-color: #0d0e12;
                color: #f8fafc;
                margin: 0;
                padding: 40px;
            }}
            .header {{
                display: flex;
                justify-content: space-between;
                align-items: center;
                border-bottom: 2px solid #2d2d2d;
                padding-bottom: 20px;
                margin-bottom: 30px;
            }}
            .logo {{
                font-size: 24px;
                font-weight: bold;
                background: linear-gradient(135deg, #a78bfa 0%, #6366f1 100%);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
            }}
            .title {{
                font-size: 20px;
                font-weight: 600;
                color: #e2e8f0;
            }}
            .meta-info {{
                color: #94a3b8;
                font-size: 14px;
            }}
            .summary-cards {{
                display: flex;
                gap: 20px;
                margin-bottom: 30px;
            }}
            .card {{
                flex: 1;
                background: #151821;
                border: 1px solid #2d2d2d;
                border-radius: 12px;
                padding: 20px;
            }}
            .card-title {{
                color: #94a3b8;
                font-size: 12px;
                text-transform: uppercase;
                margin-bottom: 8px;
                letter-spacing: 0.05em;
            }}
            .card-value {{
                font-size: 24px;
                font-weight: 700;
                color: #f8fafc;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 30px;
            }}
            th {{
                background-color: #151821;
                color: #94a3b8;
                text-align: left;
                padding: 12px;
                font-size: 12px;
                text-transform: uppercase;
                letter-spacing: 0.05em;
                border-bottom: 2px solid #2d2d2d;
            }}
            .footer {{
                text-align: center;
                color: #64748b;
                font-size: 12px;
                margin-top: 50px;
                border-top: 1px solid #2d2d2d;
                padding-top: 20px;
            }}
            @media print {{
                body {{
                    background-color: #ffffff;
                    color: #000000;
                    padding: 0;
                }}
                th {{
                    background-color: #f1f5f9;
                    color: #475569;
                    border-bottom: 2px solid #cbd5e1;
                }}
                td {{
                    color: #000000 !important;
                    border-bottom: 1px solid #e2e8f0 !important;
                }}
                .card {{
                    background: #ffffff;
                    border: 1px solid #e2e8f0;
                }}
                .card-value {{
                    color: #000000;
                }}
                .logo {{
                    -webkit-text-fill-color: #6366f1;
                }}
            }}
        </style>
    </head>
    <body onload="window.print()">
        <div class="header">
            <div>
                <div class="logo">BizPilot AI OS</div>
                <div class="title">{business.name} - Operational Report</div>
            </div>
            <div class="meta-info" style="text-align: right;">
                <div>Date: {datetime.date.today().strftime("%B %d, %Y")}</div>
                <div>Format: PDF Operational Export</div>
                <div>Status: Generated by AI Business Analyst</div>
            </div>
        </div>
        
        <div class="summary-cards">
            <div class="card">
                <div class="card-title">Total Product Skus</div>
                <div class="card-value">{len(products)}</div>
            </div>
            <div class="card">
                <div class="card-title">Total Items in Stock</div>
                <div class="card-value">{total_stock}</div>
            </div>
            <div class="card">
                <div class="card-title">Estimated Stock Asset Value</div>
                <div class="card-value">${total_val:.2f}</div>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th style="width: 12%;">SKU</th>
                    <th style="width: 30%;">Product Name</th>
                    <th style="width: 15%;">Category</th>
                    <th style="width: 10%; text-align: right;">Cost</th>
                    <th style="width: 10%; text-align: right;">Price</th>
                    <th style="width: 8%; text-align: center;">Stock</th>
                    <th style="width: 8%; text-align: center;">Safety</th>
                    <th style="width: 12%; text-align: center;">Status</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>
        
        <div class="footer">
            BizPilot AI OS — The AI Operating System for Small Businesses. Generated dynamically.
        </div>
    </body>
    </html>
    """
    return HTMLResponse(content=html_content)

@router.get("/export/excel")
def export_excel_report(report_type: str = "inventory", current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    business = db.query(Business).filter(Business.user_id == current_user.id).first()
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    # Enable gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Styled header
    ws.append(["SKU", "Product Name", "Category", "Cost ($)", "Price ($)", "Stock Level", "Safety Threshold", "Estimated Asset Value ($)"])
    
    # Apply styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo theme
    center_align = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    products = db.query(Product).filter(Product.business_id == business.id).all()
    for p in products:
        stock = p.inventory.stock_level if p.inventory else 0
        safety = p.inventory.safety_threshold if p.inventory else 10
        asset_val = stock * (p.cost or 0)
        ws.append([
            p.sku, 
            p.name, 
            p.category, 
            float(p.cost or 0), 
            float(p.price or 0), 
            int(stock), 
            int(safety),
            float(asset_val)
        ])
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Save workbook to memory
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=bizpilot_{report_type}_report.xlsx"}
    )
